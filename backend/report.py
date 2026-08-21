from schemas import (
    BirimCreate, BirimOut ,
    BantCreate , BantOut,
    PersonelCreate, PersonelOut,
    KayitCreate, KayitOut,
    EkMesaiCreate, EkMesaiOut,
    PersonelSicilNo,
    LoginRequest
    )
from sqlalchemy import func
from models import Birim , Bant , Personel , Kayit ,Ek_Mesai
import pandas as pd


from openpyxl import load_workbook
from openpyxl.styles import PatternFill


def kayit_verisi_getir(db, hedef_birim, hedef_tarih):
    sonuc = db.query(
        Kayit.sicil_no,
        Personel.ad_soyad,
        Personel.servis_no, 
        Bant.bant_adi,
        Kayit.giris_saati,
        Kayit.cikis_saati
    ).join(Personel, Kayit.sicil_no == Personel.sicil_no) \
    .join(Bant, Kayit.bant_no == Bant.bant_no) \
    .join(Birim, Bant.birim_no == Birim.birim_no
    ).filter(func.date(Kayit.giris_saati)== hedef_tarih,Birim.birim_no==hedef_birim
    ).all()
    return sonuc

def ek_mesai_verisi_getir(db, hedef_birim, hedef_tarih):
    sonuc = db.query(
        Ek_Mesai.sicil_no,
        Personel.ad_soyad,
        Personel.servis_no,
        Ek_Mesai.baslangic_saati,
        Ek_Mesai.bitis_saati
    ).join(Personel, Ek_Mesai.sicil_no == Personel.sicil_no
    ).filter(
        func.date(Ek_Mesai.baslangic_saati) == hedef_tarih,
        Ek_Mesai.birim_no == hedef_birim
    ).all()
    return sonuc

def kayit_raporu_olustur(db, hedef_birim, hedef_tarih):
    veri = kayit_verisi_getir(db, hedef_birim, hedef_tarih)
    df = pd.DataFrame(veri, columns=["sicil_no", "ad_soyad","servis_no", "bant_adi", "giris_saati", "cikis_saati"])  
    df["cikis_saati"] = df["cikis_saati"].fillna("Aktif")
    df["servis_no"] = df["servis_no"].fillna("Servis Yok")
    return df

def ek_mesai_raporu_olustur(db, hedef_birim, hedef_tarih):
    veri = ek_mesai_verisi_getir(db, hedef_birim, hedef_tarih)
    df = pd.DataFrame(veri, columns=["sicil_no", "ad_soyad","servis_no","baslangic_saati", "bitis_saati"])  
    df["bitis_saati"] = df["bitis_saati"].fillna("Aktif")
    df["servis_no"] = df["servis_no"].fillna("Servis Yok")
    return df

def excel_dosyalari_olustur(db, hedef_birim, birim_adi, hedef_tarih):
    kayit_df = kayit_raporu_olustur(db, hedef_birim, hedef_tarih)
    ek_mesai_df = ek_mesai_raporu_olustur(db, hedef_birim, hedef_tarih)

    tarih_str = hedef_tarih.strftime("%d.%m.%Y")

    kayit_dosya_adi = f"{tarih_str} {birim_adi}.xlsx"
    ek_mesai_dosya_adi = f"{tarih_str} {birim_adi}_ekMesai.xlsx"

    kayit_df.to_excel(kayit_dosya_adi, index=False)
    ek_mesai_df.to_excel(ek_mesai_dosya_adi, index=False)

    excel_bicimlendir(kayit_dosya_adi)
    excel_bicimlendir(ek_mesai_dosya_adi)
    return kayit_dosya_adi, ek_mesai_dosya_adi

def excel_bicimlendir(dosya_adi):
    wb = load_workbook(dosya_adi)
    ws = wb.active

    sari_dolgu = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # Başlık satırını sarı yap
    for hucre in ws[1]:
        hucre.fill = sari_dolgu
    # Her sütunu içeriğine göre otomatik genişlet
    for sutun in ws.columns:
        max_uzunluk = max(len(str(hucre.value)) for hucre in sutun if hucre.value is not None)
        sutun_harfi = sutun[0].column_letter
        ws.column_dimensions[sutun_harfi].width = max_uzunluk + 2

    wb.save(dosya_adi)
   
